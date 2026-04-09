"""
CarbonMonitor 中国省份航空碳排放数据处理脚本
============================================================
数据来源：https://carbonmonitor.org.cn/?by=China
目标行业：国内航空 (Domestic Aviation)
输出文件：aviation_co2_annual_province.xlsx

【年度值计算方法】
  日度数据（每年约365条/省份）→ 去掉最高值和最低值各1个 → 取均值 → ×365 → MtCO2/年

【数据说明】
  carbonmonitor.org.cn 省级数据仅提供 2019 年至今
  2013-2018 年省级航空数据该网站未收录

【使用方式】
  1. 直接运行（使用已有 CSV 数据）：
       python carbonmonitor.py

  2. 重新抓取最新网页数据（需联网）：
       python carbonmonitor.py --fetch

  3. 仅展示数据摘要，不写 Excel：
       python carbonmonitor.py --summary
============================================================
"""

import os
import re
import sys
import time
import json
import warnings
import argparse
import requests
import numpy as np
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ────────────────────────────────────────────────────────────
# 省份映射（中英文，31省份标准顺序）
# ────────────────────────────────────────────────────────────
PROVINCE_MAP_CN_EN = {
    "北京": "Beijing",      "天津": "Tianjin",       "河北": "Hebei",
    "山西": "Shanxi",       "内蒙古": "Inner Mongolia","辽宁": "Liaoning",
    "吉林": "Jilin",        "黑龙江": "Heilongjiang",  "上海": "Shanghai",
    "江苏": "Jiangsu",      "浙江": "Zhejiang",       "安徽": "Anhui",
    "福建": "Fujian",       "江西": "Jiangxi",        "山东": "Shandong",
    "河南": "Henan",        "湖北": "Hubei",          "湖南": "Hunan",
    "广东": "Guangdong",    "广西": "Guangxi",        "海南": "Hainan",
    "重庆": "Chongqing",    "四川": "Sichuan",        "贵州": "Guizhou",
    "云南": "Yunnan",       "西藏": "Tibet",          "陕西": "Shaanxi",
    "甘肃": "Gansu",        "青海": "Qinghai",        "宁夏": "Ningxia",
    "新疆": "Xinjiang",
}
EN_TO_CN = {v: k for k, v in PROVINCE_MAP_CN_EN.items()}
PROVINCE_ORDER = list(PROVINCE_MAP_CN_EN.keys())  # 标准行政区划顺序

SCRIPT_DIR = Path(__file__).parent
CSV_FILE = SCRIPT_DIR / "china_province_co2_daily.csv"
OUTPUT_FILE = SCRIPT_DIR / "aviation_co2_annual_province.xlsx"
PAGE_CACHE = SCRIPT_DIR / "raw_data" / "china_page.html"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://carbonmonitor.org.cn/?by=China",
}

# ────────────────────────────────────────────────────────────
# 核心计算函数
# ────────────────────────────────────────────────────────────

def trimmed_mean_annual(values):
    """
    去掉最高值和最低值各1个后取均值，再乘以365，得到年度排放量（MtCO2/年）。
    若数据点 ≤2 则直接取均值×365；若为空则返回 NaN。
    """
    arr = np.array([v for v in values if v is not None and not np.isnan(v)], dtype=float)
    if len(arr) == 0:
        return np.nan
    if len(arr) <= 2:
        return float(np.mean(arr)) * 365
    trimmed = np.sort(arr)[1:-1]
    return float(np.mean(trimmed)) * 365


# ────────────────────────────────────────────────────────────
# 数据读取：从 CSV（2019-2023）
# ────────────────────────────────────────────────────────────

def load_from_csv():
    """从本地 CSV 读取 2019-2023 年航空碳排放日度数据"""
    if not CSV_FILE.exists():
        print(f"  ✗ 找不到 CSV 文件: {CSV_FILE}")
        return pd.DataFrame()

    df = pd.read_csv(CSV_FILE)
    df_avi = df[df["sector_en"] == "Domestic Aviation"].copy()
    print(f"  ✓ CSV 航空记录: {len(df_avi):,} 条，年份: {sorted(df_avi['year'].unique())}")
    return df_avi


# ────────────────────────────────────────────────────────────
# 数据读取：从网页（解析 Highcharts JS 数据）
# ────────────────────────────────────────────────────────────

def fetch_page(force=False):
    """
    获取 carbonmonitor.org.cn 中国页面 HTML。
    若缓存存在且 force=False，则直接读缓存；否则联网抓取。
    """
    cache = PAGE_CACHE
    cache.parent.mkdir(parents=True, exist_ok=True)

    if cache.exists() and not force:
        print(f"  使用缓存页面: {cache}")
        with open(cache, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    url = "https://carbonmonitor.org.cn/user/data.php?by=China"
    print(f"  抓取页面: {url}")
    try:
        resp = requests.get(url, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        content = resp.text
        with open(cache, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"  ✓ 页面已缓存 ({len(content)//1024} KB): {cache}")
        return content
    except Exception as e:
        print(f"  ✗ 抓取失败: {e}")
        return ""


def parse_page_aviation(content):
    """
    从 Highcharts 页面 JS 中解析 Domestic Aviation 日度数据。
    返回 DataFrame，列：province_en, year, day_of_year, value。
    """
    if not content:
        return pd.DataFrame()

    soup = BeautifulSoup(content, "html.parser")
    scripts = soup.find_all("script")
    # 选最大的 script（包含所有图表数据）
    script_text = max((s.get_text() for s in scripts), key=len, default="")
    print(f"  数据脚本大小: {len(script_text)//1024} KB")

    # 按 Highcharts.chart 块切割
    chart_blocks = re.split(r"(?=Highcharts\.chart\s*\()", script_text)

    rows = []
    found_provinces = set()
    for block in chart_blocks:
        m = re.search(r'text\s*:\s*"([^"]+)\s*-\s*Aviation"', block)
        if not m:
            continue
        prov_en = m.group(1).strip()
        if prov_en not in EN_TO_CN:
            continue
        found_provinces.add(prov_en)

        s_match = re.search(r"series\s*:\s*(\[.*?\])\s*[,}]", block, re.DOTALL)
        if not s_match:
            continue
        series_text = s_match.group(1)

        for yb in re.finditer(r'\{"name":(\d{4}),.*?"data":\[([\d.,\s]+)\]', series_text, re.DOTALL):
            year = int(yb.group(1))
            vals = [float(v) for v in re.findall(r"[\d.]+", yb.group(2))]
            for day_idx, val in enumerate(vals):
                rows.append({"province_en": prov_en, "year": year, "day_of_year": day_idx + 1, "value": val})

    df = pd.DataFrame(rows)
    if len(df) > 0:
        print(f"  ✓ 页面解析: {len(df):,} 条，{len(found_provinces)} 省份，年份: {sorted(df['year'].unique())}")
    else:
        print("  ✗ 页面未解析到有效数据")
    return df


# ────────────────────────────────────────────────────────────
# 汇总：计算年度排放量
# ────────────────────────────────────────────────────────────

def build_annual(df_csv, df_page):
    """
    合并 CSV（2019-2023）与页面数据（补充 2024+），计算年度排放量。
    CSV 数据优先（精度更高），页面数据补充新年份。
    返回 DataFrame，列：province_cn, province_en, year, annual_mtco2。
    """
    rows = []
    csv_years = set()

    # ── CSV 数据（2019-2023）──
    if len(df_csv) > 0:
        for (prov_en, prov_cn, yr), grp in df_csv.groupby(["province_en", "province_cn", "year"]):
            val = trimmed_mean_annual(grp["value_mtco2_per_day"].values)
            rows.append({"province_cn": prov_cn, "province_en": prov_en, "year": int(yr), "annual_mtco2": val})
            csv_years.add(int(yr))

    # ── 页面数据（补充新年份）──
    if len(df_page) > 0:
        new_years = [y for y in df_page["year"].unique() if y not in csv_years]
        if new_years:
            print(f"  页面补充新年份: {sorted(new_years)}")
            for (prov_en, yr), grp in df_page[df_page["year"].isin(new_years)].groupby(["province_en", "year"]):
                prov_cn = EN_TO_CN.get(prov_en, prov_en)
                val = trimmed_mean_annual(grp["value"].values)
                rows.append({"province_cn": prov_cn, "province_en": prov_en, "year": int(yr), "annual_mtco2": val})
        else:
            print("  页面未提供新年份数据")

    df = pd.DataFrame(rows).drop_duplicates(subset=["province_en", "year"])
    df["_sort"] = df["province_cn"].apply(lambda x: PROVINCE_ORDER.index(x) if x in PROVINCE_ORDER else 99)
    df = df.sort_values(["year", "_sort"]).drop(columns=["_sort"]).reset_index(drop=True)
    return df


# ────────────────────────────────────────────────────────────
# 透视矩阵
# ────────────────────────────────────────────────────────────

def build_pivot(df_annual):
    """生成 年份×省份 透视矩阵"""
    pivot = df_annual.pivot_table(index="year", columns="province_cn", values="annual_mtco2", aggfunc="first")
    # 按标准省份顺序排列列
    ordered_cols = [c for c in PROVINCE_ORDER if c in pivot.columns]
    pivot = pivot[ordered_cols + [c for c in pivot.columns if c not in ordered_cols]]
    pivot.index.name = "年份"
    pivot.columns.name = "省份"
    return pivot


# ────────────────────────────────────────────────────────────
# 写出 Excel
# ────────────────────────────────────────────────────────────

def write_excel(pivot, df_annual, output_path):
    """生成带格式的 Excel 文件（Sheet1: 矩阵，Sheet2: 明细）"""
    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF", size=10)
    year_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    year_font = Font(bold=True, size=10)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Sheet1: 年份×省份矩阵 ──
        pivot.round(4).to_excel(writer, sheet_name="航空CO2矩阵(年份×省份)")
        ws1 = writer.sheets["航空CO2矩阵(年份×省份)"]

        for cell in ws1[1]:
            if cell.value is not None:
                cell.font = white_bold
                cell.fill = blue_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws1.iter_rows(min_row=2):
            row[0].font = year_font
            row[0].fill = year_fill
            row[0].alignment = Alignment(horizontal="center")
            for cell in row[1:]:
                cell.alignment = Alignment(horizontal="center")
                if cell.value is not None:
                    cell.number_format = "0.0000"

        ws1.row_dimensions[1].height = 40
        ws1.column_dimensions["A"].width = 8
        for col_idx in range(2, ws1.max_column + 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 8
        ws1.freeze_panes = "B2"

        # ── Sheet2: 明细数据 ──
        detail = df_annual[["year", "province_cn", "province_en", "annual_mtco2"]].copy()
        detail.columns = ["年份", "省份(中文)", "省份(英文)", "年度CO2排放量(MtCO2/年)"]
        detail.to_excel(writer, sheet_name="明细数据", index=False)
        ws2 = writer.sheets["明细数据"]
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")
        for w, col in zip([8, 12, 18, 24], ["A", "B", "C", "D"]):
            ws2.column_dimensions[col].width = w


# ────────────────────────────────────────────────────────────
# 打印摘要
# ────────────────────────────────────────────────────────────

def print_summary(pivot):
    years = pivot.index.tolist()
    print(f"\n{'='*60}")
    print(f"  矩阵形状: {pivot.shape[0]} 年份 × {pivot.shape[1]} 省份")
    print(f"  年份列表: {years}")
    print(f"\n  各年全国平均排放量 (MtCO2/年/省):")
    for yr, val in pivot.mean(axis=1).round(4).items():
        print(f"    {yr}: {val:.4f}")
    print(f"\n  排名前5省份 (2023年或最新年):")
    latest_year = max(years)
    top5 = pivot.loc[latest_year].sort_values(ascending=False).head(5)
    for prov, val in top5.items():
        print(f"    {prov}: {val:.4f} MtCO2")
    print(f"{'='*60}")


# ────────────────────────────────────────────────────────────
# 主流程
# ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="国内航空碳排放数据处理工具")
    parser.add_argument("--fetch", action="store_true", help="强制重新抓取最新网页数据")
    parser.add_argument("--summary", action="store_true", help="仅展示摘要，不写 Excel")
    args = parser.parse_args()

    print("=" * 60)
    print("  CarbonMonitor · 国内航空碳排放年度矩阵")
    print("  行业: Domestic Aviation | 单位: MtCO2/年")
    print("=" * 60)

    # 1. 读取 CSV
    print("\n[1/4] 读取本地 CSV 数据 (2019-2023)...")
    df_csv = load_from_csv()

    # 2. 解析网页（补充 2024+ 或强制更新）
    print("\n[2/4] 解析网页数据 (补充最新年份)...")
    page_content = fetch_page(force=args.fetch)
    df_page = parse_page_aviation(page_content)

    # 3. 计算年度值
    print("\n[3/4] 计算年度排放量 (截尾均值×365)...")
    df_annual = build_annual(df_csv, df_page)
    pivot = build_pivot(df_annual)

    years_available = sorted(df_annual["year"].unique())
    print(f"  数据覆盖: {len(years_available)} 年 × {df_annual['province_en'].nunique()} 省份")
    print(f"  年份列表: {years_available}")

    print_summary(pivot)

    # 4. 写出 Excel
    if not args.summary:
        print(f"\n[4/4] 写出 Excel: {OUTPUT_FILE}")
        write_excel(pivot, df_annual, OUTPUT_FILE)
        print(f"  ✓ 已保存: {OUTPUT_FILE}")
        print(f"    Sheet1 → 航空CO2矩阵(年份×省份)")
        print(f"    Sheet2 → 明细数据")
    else:
        print("\n[4/4] --summary 模式，跳过 Excel 输出")

    print("\n【注意】carbonmonitor.org.cn 省级数据仅提供 2019 年至今")
    print("        如需 2013-2018 年数据，请参考中国民航局或 CEADs 数据库")


if __name__ == "__main__":
    main()
